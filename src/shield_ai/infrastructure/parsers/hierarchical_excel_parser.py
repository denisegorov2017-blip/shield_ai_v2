import logging
from typing import List, Union, Tuple
from pathlib import Path
from datetime import datetime
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from src.shield_ai.infrastructure.parsers.dto import FlatRecord, RowType, ParsingContext


def recognize_row_type(row_data: list) -> RowType:
    """
    Определяет тип строки в Excel-отчёте на основе более строгих правил.
    """
    if all(value is None or str(value).strip() == '' for value in row_data):
        return RowType.EMPTY

    str_values = [str(v).strip().lower() if v is not None else '' for v in row_data]
    
    # 1. Проверка на TOTAL
    if any("итого" in s for s in str_values):
        return RowType.TOTAL

    # 2. Проверка на HEADER
    header_keywords = ["номенклатура", "начальный остаток", "приход", "расход", "конечный остаток"]
    matches = sum(1 for keyword in header_keywords if any(keyword in s for s in str_values))
    if matches >= 3:  # Если найдено 3 или более ключевых слова, это заголовок
        return RowType.HEADER

    # 3. Проверка на DATA
    # Строка данных должна содержать числовые значения в колонках количества
    has_any_qty = False
    # Check columns for qty_begin, qty_in, qty_out, qty_end (indices 7, 8, 9, 10)
    for i in range(7, 11):  
        if i < len(row_data) and row_data[i] is not None and str(row_data[i]).strip() != '':
            try:
                Decimal(str(row_data[i]))
                has_any_qty = True
                break
            except InvalidOperation:
                continue
    if has_any_qty:
        return RowType.DATA

    # 4. Проверка на GROUP_HEADER
    # Если в строке есть текст, но нет данных о количестве, это заголовок группы/товара
    non_empty_cells = [s for s in str_values if s]
    if 0 < len(non_empty_cells) <= 3: # Обычно в заголовке группы 1-3 непустых ячеек
        is_group = True
        for cell in non_empty_cells:
            try:
                # If it can be converted to a float, it's probably not a group header
                float(cell)
                is_group = False
                break
            except ValueError:
                pass
        if is_group:
            return RowType.GROUP_HEADER

    return RowType.UNDEFINED

class HierarchicalExcelParser:
    """
    Парсер для иерархических Excel-файлов, преобразующий их в плоскую структуру FlatRecord.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def parse(self, file_path: Union[str, Path]) -> Tuple[List[FlatRecord], List[dict]]:
        """
        Парсит иерархический Excel-файл и возвращает список плоских записей и логи ошибок.
        
        Returns:
            Tuple[List[FlatRecord], List[dict]]: Кортеж из списка плоских записей и списка ошибок
        """
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        
        if ws is None:
            return [], []
        
        context = ParsingContext()
        results: List[FlatRecord] = []
        error_logs: List[dict] = []
        
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            row_data = [cell.value for cell in row]
            row_type = recognize_row_type(row_data)
            
            if row_type in [RowType.HEADER, RowType.TOTAL, RowType.EMPTY, RowType.UNDEFINED]:
                if row_type == RowType.TOTAL:
                    context.product = None # Сброс контекста продукта после итога
                continue

            elif row_type == RowType.GROUP_HEADER:
                # Обновляем контекст на основе заголовка группы
                for cell_value in row_data:
                    if cell_value and isinstance(cell_value, str) and cell_value.strip():
                        # Первая непустая текстовая ячейка - это продукт/группа
                        context.product = cell_value.strip()
                        break
                continue

            elif row_type == RowType.DATA:
                try:
                    product = context.product or ""
                    if not product:
                        error_msg = f"Skipping DATA row {row_idx} due to missing product context."
                        self.logger.warning(error_msg)
                        error_logs.append({
                            'row': row_idx,
                            'type': 'warning',
                            'message': error_msg,
                            'data': str(row_data)
                        })
                        continue

                    # Извлекаем остальные данные
                    batch_code = str(row_data[3]).strip() if len(row_data) > 3 and row_data[3] is not None else ""
                    
                    batch_date = None
                    if len(row_data) > 4 and row_data[4] is not None:
                        if isinstance(row_data[4], datetime):
                            batch_date = row_data[4]
                        else:
                            try:
                                batch_date = datetime.strptime(str(row_data[4]), '%d.%m.%Y %H:%M:%S')
                            except (ValueError, TypeError):
                                try:
                                    batch_date = datetime.strptime(str(row_data[4]), '%Y-%m-%d')
                                except (ValueError, TypeError):
                                    self.logger.warning(f"Could not parse batch_date '{row_data[4]}' at row {row_idx}.")
                                    error_logs.append({
                                        'row': row_idx,
                                        'type': 'warning',
                                        'message': f"Could not parse batch_date '{row_data[4]}'",
                                        'data': str(row_data[4])
                                    })
                                    pass
                    
                    doc_type = str(row_data[5]).strip() if len(row_data) > 5 and row_data[5] is not None else ""
                    
                    doc_date = None
                    if len(row_data) > 6 and row_data[6] is not None:
                        if isinstance(row_data[6], datetime):
                            doc_date = row_data[6]
                        else:
                            try:
                                doc_date = datetime.strptime(str(row_data[6]), '%d.%m.%Y %H:%M:%S')
                            except (ValueError, TypeError):
                                try:
                                    doc_date = datetime.strptime(str(row_data[6]), '%Y-%m-%d')
                                except (ValueError, TypeError):
                                    self.logger.warning(f"Could not parse doc_date '{row_data[6]}' at row {row_idx}.")
                                    error_logs.append({
                                        'row': row_idx,
                                        'type': 'warning',
                                        'message': f"Could not parse doc_date '{row_data[6]}'",
                                        'data': str(row_data[6])
                                    })
                                    pass

                    def to_decimal(value, default=Decimal('0')):
                        if value is None or str(value).strip() == '':
                            return default
                        try:
                            return Decimal(str(value))
                        except InvalidOperation:
                            return default

                    qty_begin = to_decimal(row_data[7] if len(row_data) > 7 else None)
                    qty_in = to_decimal(row_data[8] if len(row_data) > 8 else None)
                    qty_out = to_decimal(row_data[9] if len(row_data) > 9 else None)
                    qty_end = to_decimal(row_data[10] if len(row_data) > 10 else None)
                    
                    unit = str(row_data[11]).strip() if len(row_data) > 11 and row_data[11] is not None else "шт"
                    comment = str(row_data[12]).strip() if len(row_data) > 12 and row_data[12] is not None else None

                    final_batch_date = batch_date or doc_date or context.batch_date or datetime.now()
                    final_doc_date = doc_date or final_batch_date

                    flat_record = FlatRecord(
                        warehouse=context.warehouse or "Не определен",
                        group=context.group or "",
                        product=product,
                        batch_code=batch_code,
                        batch_date=final_batch_date,
                        doc_type=doc_type,
                        doc_date=final_doc_date,
                        qty_begin=qty_begin,
                        qty_in=qty_in,
                        qty_out=qty_out,
                        qty_end=qty_end,
                        unit=unit,
                        comment=comment
                    )
                    results.append(flat_record)

                except (IndexError, ValueError, TypeError) as e:
                    error_msg = f"Error processing DATA row {row_idx}: {e}. Row data: {row_data}"
                    self.logger.error(error_msg)
                    error_logs.append({
                        'row': row_idx,
                        'type': 'error',
                        'message': str(e),
                        'data': str(row_data)
                    })
                    continue
        
        return results, error_logs
